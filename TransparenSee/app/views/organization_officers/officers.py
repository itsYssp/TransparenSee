from django.utils import timezone
from django.shortcuts import render
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView, DetailView
from ...forms import *
from ..mixins import *
from django.db.models import Sum, Q, Count 
from ...blockchain import verify_report_hash
from itertools import groupby  
import openpyxl
from django.contrib import messages
from django.shortcuts import redirect
from django.views import View
from django.utils.crypto import get_random_string
from django.db import transaction

class ApproveReportView(RoleRequireMixin, TemplateView):
    role_required = ['auditor', 'president', 'adviser', 'co_adviser' ]

    def post(self, request, pk):
        report = get_object_or_404(FinancialReport, pk=pk)
        action = request.POST.get('action')
        remarks = request.POST.get('remarks', '')
        user = request.user

        if action == 'approve':
            if user.role == 'auditor' and report.status == 'pending_auditor':
                report.auditor_approved_by = user
                report.auditor_approved_at = timezone.now()
                report.auditor_remarks = remarks
                report.status = 'pending_president'
                report.save()
                ReportApprovalLog.objects.create(
                    report=report, action_by=user,
                    action='approved', remarks=remarks
                )
                messages.success(request, 'Report approved. Sent to President.')

            elif user.role == 'president' and report.status == 'pending_president':
                report.president_approved_by = user
                report.president_approved_at = timezone.now()
                report.president_remarks = remarks
                report.status = 'pending_co_adviser'
                report.save()
                ReportApprovalLog.objects.create(
                    report=report, action_by=user,
                    action='approved', remarks=remarks
                )
                messages.success(request, 'Report approved. Sent to Co-Adviser.')

            elif user.role == 'co_adviser' and report.status == 'pending_co_adviser':
                report.co_adviser_approved_by= user
                report.co_adviser_approved_at = timezone.now()
                report.co_adviser_remarks = remarks
                report.status = 'pending_adviser'
                report.save()
                ReportApprovalLog.objects.create(
                    report=report, action_by=user,
                    action='approved', remarks=remarks
                )
                messages.success(request, 'Report approved. Sent to Adviser.')

            elif user.role == 'adviser' and report.status == 'pending_adviser':
                report.adviser_approved_by = user
                report.adviser_approved_at = timezone.now()
                report.adviser_remarks = remarks
                report.status = 'approved'
                report.save()
                
                entities = report.entries.all()
                total_income = entities.filter(entry_type='income').aggregate(t=Sum('amount'))['t'] or 0
                total_expense = entities.filter(entry_type='expense').aggregate(t=Sum('amount'))['t'] or 0
                net = total_income - total_expense

                org = report.organization
                org.balance = (org.balance or 0) + net
                org.save(update_fields=['balance'])

                ReportApprovalLog.objects.create(
                    report=report, action_by=user,
                    action='approved', remarks=remarks
                )
                messages.success(request, 'Report fully approved and ready for blockchain.')

        elif action == 'reject':
            report.status = 'rejected'
            report.save()
            ReportApprovalLog.objects.create(
                report=report, action_by=user,
                action='rejected', remarks=remarks
            )
            messages.error(request, 'Report rejected.')

        return redirect('report_detail', pk=pk)

class ReportListView(RoleRequireMixin, ListView):
    role_required = ['treasurer', 'auditor', 'vice_president', 'president', 'adviser', 'co_adviser']
    model = FinancialReport
    template_name = 'app/officer/treasurer/report_list.html'
    paginate_by = 10

    def get_queryset(self):
        user = self.request.user
        qs = super().get_queryset().select_related(
            'organization', 'created_by', 'academic_year'
        ).prefetch_related('entries')

        if hasattr(user, 'officer'):
            qs = qs.filter(organization=user.officer.organization)
        elif hasattr(user, 'adviser'):
            qs = qs.filter(organization=user.adviser.organization)
        elif hasattr(user, 'campus_admin'):
            qs = qs  


        search = self.request.GET.get('search', '').strip()
        if search:
            qs = qs.filter(title__icontains=search)


        status = self.request.GET.get('status')
        if status   :
            qs = qs.filter(status=status)

        qs = qs.annotate(
        total_income=Sum('entries__amount', filter=Q(entries__entry_type='income')),
        total_expense=Sum('entries__amount', filter=Q(entries__entry_type='expense')),
        income_count=Count('entries', filter=Q(entries__entry_type='income')),
        expense_count=Count('entries', filter=Q(entries__entry_type='expense')),
        )

        
        qs = qs.order_by('-created_at')

        return qs

    def get_organization(self, user):
        if hasattr(user, 'officer'):
            return user.officer.organization
        elif hasattr(user, 'adviser'):
            return user.adviser.organization
        elif hasattr(user, 'co_adviser'):
            return user.adviser.organization
        return None

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        org = self.get_organization(user)

        role_templates = {
            "treasurer": "app/officer/treasurer/sidebar.html",
            "auditor": "app/officer/auditor/sidebar.html",
            "president": "app/officer/president/sidebar.html",
            "vice_president": "app/officer/president/sidebar.html",
            "adviser": "app/adviser/sidebar.html",
            "co_adviser": "app/adviser/sidebar.html",
        }

        context['base_template'] = role_templates.get(user.role, "app/base.html")

        context['reports'] = context['object_list']

        context['rejected_reports'] = FinancialReport.objects.filter(
            organization=org, status='rejected'
        ).annotate(
            total_income=Sum('entries__amount', filter=Q(entries__entry_type='income')),
            total_expense=Sum('entries__amount', filter=Q(entries__entry_type='expense')),
            income_count=Count('entries', filter=Q(entries__entry_type='income')),
            expense_count=Count('entries', filter=Q(entries__entry_type='expense')),
        )

        context['status_choices'] = FinancialReport.STATUS_CHOICES
        return context


class ReportDetailView(RoleRequireMixin, DetailView):
    model = FinancialReport
    template_name = 'app/officer/report_details.html'
    context_object_name = 'report'
    role_required = ['treasurer', 'auditor', 'president', 'adviser', 'co_adviser', 'vice_president']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        report = self.get_object()
        user = self.request.user


        entries = report.entries.order_by('date', 'category', 'order')

        grouped = {}
        for entry in entries:
            key = entry.date
            if key not in grouped:
                grouped[key] = {
                    'date': key,
                    'entries': [],
                    'income_subtotal': 0,
                    'expense_subtotal': 0,
                }
            grouped[key]['entries'].append(entry)
            if entry.entry_type == 'income':
                grouped[key]['income_subtotal'] += entry.amount or 0
            else:
                grouped[key]['expense_subtotal'] += entry.amount or 0

        grouped_entries = []
        for group in grouped.values():
            entries_with_meta = []
            for category, cat_entries in groupby(group['entries'], key=lambda e: e.category):
                cat_list = list(cat_entries)
                for i, entry in enumerate(cat_list):
                    entries_with_meta.append({
                        'entry':            entry,
                        'show_category':    i == 0,
                        'category_rowspan': len(cat_list) if i == 0 else None,
                    })

            grouped_entries.append({
                'date':             group['date'],
                'entries':          entries_with_meta,
                'income_subtotal':  group['income_subtotal'] or None,
                'expense_subtotal': group['expense_subtotal'] or None,
            })

        role_templates = {
            "treasurer":  "app/officer/treasurer/sidebar.html",
            "auditor":    "app/officer/auditor/sidebar.html",
            "president":  "app/officer/president/sidebar.html",
            "vice_president":  "app/officer/president/sidebar.html",
            "adviser":    "app/adviser/sidebar.html",
            "co_adviser": "app/adviser/sidebar.html",
        }

        total_income  = entries.filter(entry_type='income').aggregate(t=Sum('amount'))['t'] or 0
        total_expense = entries.filter(entry_type='expense').aggregate(t=Sum('amount'))['t'] or 0
        receipt_entries = entries.exclude(receipt_image='').exclude(receipt_image__isnull=True)

        context['base_template']       = role_templates.get(user.role, "app/base.html")
        context['grouped_entries']     = grouped_entries
        context['approval_logs']       = report.approval_logs.order_by('created_at')
        context['total_income']        = total_income
        context['total_expense']       = total_expense
        context['net_total']           = total_income - total_expense
        context['receipt_entries']     = receipt_entries
        context['blockchain_verified'] = verify_report_hash(report) if report.blockchain_hash else None

        return context

class ChatView(RoleRequireMixin, TemplateView):
    template_name = 'app/officer/officer_chat.html'
    role_required = ["treasurer", "auditor", "secretary" ,"president","vice_president", "adviser", "head", "co_adviser", "campus_admin", "student"]

    def get_organization(self, user):
        if hasattr(user, 'officer'):
            return user.officer.organization
        elif hasattr(user, 'adviser'):
            return user.adviser.organization
        elif hasattr(user, 'student'):
            return user.student.organization
        elif hasattr(user, 'co_adviser'):
            return user.co_adviser.organization 
        return None

    def can_post_global_announcement(self, user):
        return user.role in ['head', 'campus_admin']

    def can_post_org_announcement(self, user):
        return user.role in ['treasurer', 'auditor', 'president', 'adviser', 'co_adviser', 'secretary']

    def is_student(self, user):
        return user.role == 'student'

    def get_user_department(self, user):
        if hasattr(user, 'heads'):
            return user.heads.get_department_display()
        elif hasattr(user, 'campus_admin'):
            return user.campus_admin.department
        elif hasattr(user, 'adviser'):
            return user.adviser.get_department_display()
        elif hasattr(user, 'co_adviser'):
            return user.co_adviser.get_department_display()
        return ''

    def serialize_message(self, msg, current_user):
        organization_name = ''
        if msg.user.role in ['treasurer', 'auditor', 'president', 'secretary'] and hasattr(msg.user, 'officer'):
            organization_name = msg.user.officer.organization.name
        elif msg.user.role in ['adviser', 'co_adviser'] and hasattr(msg.user, 'adviser'):
            organization_name = msg.user.adviser.organization.name

        return {
            "id": msg.id,
            "is_self": msg.user_id == current_user.id,
            "author": msg.user.get_full_name() or msg.user.username,
            "role": msg.user.get_role_display(),
            "organization": organization_name,
            "message": msg.message,
            "time": timezone.localtime(msg.createdAt).strftime("%H:%M"),
            "avatar": msg.user.profile_image.url if getattr(msg.user, 'profile_image', None) else '',
        }

    def serialize_org_announcement(self, announcement):
        organization_name = announcement.organization.name if announcement.organization else 'Organization'
        logo_url = ''
        if getattr(announcement.organization, 'logo', None):
            try:
                logo_url = announcement.organization.logo.url
            except ValueError:
                logo_url = ''

        return {
            "id": announcement.id,
            "scope": "organization",
            "organization": organization_name,
            "author": announcement.author.get_full_name() or announcement.author.username,
            "message": announcement.message,
            "time": timezone.localtime(announcement.createdAt).strftime("%b %d, %Y %H:%M"),
            "logo": logo_url,
        }

    def serialize_global_announcement(self, announcement):
        return {
            "id": announcement.id,
            "scope": "global",
            "author": announcement.author.get_full_name() or announcement.author.username,
            "role": announcement.author.get_role_display(),
            "department": self.get_user_department(announcement.author),
            "message": announcement.message,
            "time": timezone.localtime(announcement.createdAt).strftime("%b %d, %Y %H:%M"),
        }

    def get_feed_payload(self, user):
        org = self.get_organization(user)
        student = self.is_student(user)

        # Students only see their org's announcements — no global chat, no global announcements
        if student:
            if org:
                announcements = OrganizationAnnouncement.objects.filter(
                    organization=org
                ).select_related("author", "organization").order_by("-createdAt")[:20]
            else:
                announcements = OrganizationAnnouncement.objects.none()

            return {
                "global_messages": [],
                "organization_announcements": [self.serialize_org_announcement(item) for item in announcements],
                "global_announcements": [],
                "permissions": {
                    "can_post_global_announcement": False,
                    "can_post_org_announcement": False,  # students read-only
                }
            }

        # Officers / advisers / heads / campus_admin — full feed
        global_messages = GlobalChat.objects.select_related("user").order_by("createdAt")[:50]
        global_announcements = GlobalAnnouncement.objects.select_related("author").order_by("-createdAt")[:20]

        if org:
            announcements = OrganizationAnnouncement.objects.filter(
                organization=org
            ).select_related("author", "organization").order_by("-createdAt")[:20]
        else:
            announcements = OrganizationAnnouncement.objects.none()

        return {
            "global_messages": [self.serialize_message(msg, user) for msg in global_messages],
            "organization_announcements": [self.serialize_org_announcement(item) for item in announcements],
            "global_announcements": [self.serialize_global_announcement(item) for item in global_announcements],
            "permissions": {
                "can_post_global_announcement": self.can_post_global_announcement(user),
                "can_post_org_announcement": bool(org and self.can_post_org_announcement(user)),
            }
        }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        org = self.get_organization(user)
        student = self.is_student(user)

        context['user_role'] = user.role
        context["is_student"] = student
        context["can_post_global_announcement"] = self.can_post_global_announcement(user)
        context["can_post_org_announcement"] = bool(org and self.can_post_org_announcement(user))

        if student:
            # Students only see their own org's announcements
            context["global_messages"] = GlobalChat.objects.none()
            context["global_announcements"] = GlobalAnnouncement.objects.none()
            if org:
                context["announcements"] = OrganizationAnnouncement.objects.filter(
                    organization=org
                ).select_related("author").order_by("createdAt")
            else:
                context["announcements"] = OrganizationAnnouncement.objects.none()
        else:
            # Officers and above see everything
            context["global_messages"] = GlobalChat.objects.select_related(
                "user"
            ).order_by("createdAt")[:50]
            context["global_announcements"] = GlobalAnnouncement.objects.select_related(
                "author"
            ).order_by("-createdAt")[:20]
            if org:
                context["announcements"] = OrganizationAnnouncement.objects.filter(
                    organization=org
                ).select_related("author").order_by("createdAt")
            else:
                context["announcements"] = OrganizationAnnouncement.objects.none()

        role_templates = {
            "treasurer":    "app/officer/treasurer/sidebar.html",
            "student":      "app/student/sidebar.html",
            "auditor":      "app/officer/auditor/sidebar.html",
            "secretary":      "app/officer/secretary/sidebar.html",
            "president":    "app/officer/president/sidebar.html",
            "vice_president":    "app/officer/president/sidebar.html",
            "adviser":      "app/adviser/sidebar.html",
            "co_adviser":   "app/adviser/sidebar.html",
            "head":         "app/heads/sidebar.html",
            "campus_admin": "app/campus_admin/sidebar.html",
        }
        context["base_template"] = role_templates.get(user.role, "app/base.html")
        context["chat_form"] = GlobalChatForm()
        context["announcement_form"] = AnnouncementForm()
        context["global_announcement_form"] = GlobalAnnouncementForm()
        return context

    def post(self, request, *args, **kwargs):
        tab = request.POST.get("type", "global_chat")
        user = request.user
        org = self.get_organization(user)

        if tab == "global_chat":
            form = GlobalChatForm(request.POST)
            if form.is_valid():
                chat = form.save(commit=False)
                chat.user = user
                chat.save()
            else:
                return JsonResponse({"ok": False, "errors": form.errors}, status=400)

        elif tab == "organization_announcement":
            if not org or not self.can_post_org_announcement(user):
                return JsonResponse({"ok": False, "errors": {"message": ["Not allowed to post organization announcements."]}}, status=403)
            form = AnnouncementForm(request.POST)
            if form.is_valid():
                ann = form.save(commit=False)
                ann.author = user
                ann.organization = org
                ann.save()
            else:
                return JsonResponse({"ok": False, "errors": form.errors}, status=400)

        elif tab == "global_announcement":
            if not self.can_post_global_announcement(user):
                return JsonResponse({"ok": False, "errors": {"message": ["Not allowed to post global announcements."]}}, status=403)
            form = GlobalAnnouncementForm(request.POST)
            if form.is_valid():
                ann = form.save(commit=False)
                ann.author = user
                ann.save()
            else:
                return JsonResponse({"ok": False, "errors": form.errors}, status=400)

        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"ok": True, "feed": self.get_feed_payload(user)})

        return redirect(f"{request.path}?type={tab}")


class ChatFeedView(RoleRequireMixin, TemplateView):
    role_required = ["treasurer", "auditor", "president", "adviser", "head", "co_adviser", "campus_admin", "student","secretary"]

    def get(self, request, *args, **kwargs):
        payload = ChatView().get_feed_payload(request.user)
        return JsonResponse(payload)


class OrgPublicProfileView(DetailView):
    model = Organization
    template_name = "app/officer/org_profile.html"
    context_object_name = "org"

    role_template = {
        "treasurer":  "app/officer/treasurer/sidebar.html",
        "auditor":    "app/officer/auditor/sidebar.html",
        "secretary":    "app/officer/secretary/sidebar.html",
        "president":  "app/officer/president/sidebar.html",
        "vice_president":  "app/officer/president/sidebar.html",
        "adviser":    "app/adviser/sidebar.html",
        "co_adviser": "app/adviser/sidebar.html",
        "head":       "app/heads/sidebar.html",
        "student":    "app/student/sidebar.html",
    }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        org = self.object
        user = self.request.user

        officers = Officer.objects.filter(organization=org).select_related("user")
        advisers = Adviser.objects.filter(organization=org).select_related("user")
        announcements = OrganizationAnnouncement.objects.filter(
            organization=org
        ).select_related("author")[:5]

        context['base_template'] = self.role_template.get(user.role, 'app/base.html')
        context["org_category"] = org.category
        context["officers"] = officers
        context["advisers"] = advisers
        context["total_officers"] = officers.count()
        context["total_advisers"] = advisers.count()
        context["announcements"] = announcements
        return context

class ProductListView(ListView, RoleRequireMixin):
    model = Product
    template_name = 'app/officer/product_list.html'
    context_object_name = 'products'
    role_required = ["treasurer", "auditor", "president", "adviser", "co_adviser", 'vice_president','secretary']
    paginate_by = 10

    role_templates = {
        'treasurer' : 'app/officer/treasurer/sidebar.html',
        'auditor' : 'app/officer/auditor/sidebar.html',
        "secretary":    "app/officer/secretary/sidebar.html",
        'president' : 'app/officer/president/sidebar.html',
        'vice_president' : 'app/officer/president/sidebar.html',
        'adviser' : 'app/adviser/sidebar.html',
        'co_adviser' : 'app/adviser/sidebar.html',
    }

    def get_organization(self, user):
        if hasattr(user, 'officer'):
            return self.request.user.officer.organization
        elif hasattr(user, 'adviser'):
            return self.request.user.adviser.organization
        elif hasattr(user, 'co_adviser'):
            return self.request.user.adviser.organization
        return None
        

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        org = self.get_organization(user)
        context["base_template"] = self.role_templates.get(user.role,'app/base.html')
        context["products"] = Product.objects.filter(organization=org).annotate(
            variant_count = Count('variants')
        )
        return context
    
class BlockchainRecordsView(RoleRequireMixin, TemplateView):
    template_name = 'app/blockchain_records.html'
    role_required = ["treasurer","auditor", "president", "vice_president", "adviser", "co_adviser","head", "campus_admin", "admin", 'secretary' ] 

    def get_organization(self, user):
        if hasattr(user, 'officer'):
            return user.officer.organization
        elif hasattr(user, 'adviser'):
            return user.adviser.organization
        elif hasattr(user, 'co_adviser'):
            return user.adviser.organization
        return None
    
    role_templates = {
        "treasurer": "app/officer/treasurer/sidebar.html",
        "auditor": "app/officer/auditor/sidebar.html",
        "president": "app/officer/president/sidebar.html",
        "vice_president": "app/officer/president/sidebar.html",
        "adviser": "app/adviser/sidebar.html",
        "co_adviser": "app/adviser/sidebar.html",
        "head": "app/heads/sidebar.html",
        "campus_admin": "app/campus_admin/sidebar.html",
        "admin": "app/superadmin/sidebar.html",
        "secretary":    "app/officer/secretary/sidebar.html",

    }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        context["base_template"] = self.role_templates.get(user.role, 'app/base.html')
        return context
    

    def get(self, request):
        context = self.get_context_data()
        user = self.request.user
        org = self.get_organization(user)

        reports = FinancialReport.objects.filter( 
                status='on_blockchain'
        ).prefetch_related('entries').order_by('-blockchain_recorded_at')
        
        if org:
            reports = FinancialReport.objects.filter(
                organization=org,
                status='on_blockchain',
            ).prefetch_related('entries').order_by('-blockchain_recorded_at')
            

        for report in reports:
            report.total_income  = report.entries.filter(entry_type='income').aggregate(t=Sum('amount'))['t'] or 0
            report.total_expense = report.entries.filter(entry_type='expense').aggregate(t=Sum('amount'))['t'] or 0
            report.net           = report.total_income - report.total_expense

        total_income  = sum(r.total_income for r in reports)
        total_expense = sum(r.total_expense for r in reports)

        context.update({
            'reports':        reports,
            'total_income':   total_income,
            'total_expense':  total_expense,
            'academic_years': AcademicYear.objects.order_by('-academic_year'),
        })

        return render(request, self.template_name, context)
    
class MembersView(RoleRequireMixin, TemplateView):
    template_name = "app/officer/members.html"
    role_required = ['president', 'treasurer', 'auditor', 'adviser', 'co_adviser', 'vice_president', 'secretary']

    def get_organization(self):
        user = self.request.user
        if hasattr(user, 'officer'):
            return user.officer.organization
        if hasattr(user, 'adviser'):
            return user.adviser.organization
        return None
    
    role_templates = {
        'treasurer': 'app/officer/treasurer/sidebar.html',
        'auditor': 'app/officer/auditor/sidebar.html',
        'president': 'app/officer/president/sidebar.html',
        'vice_president': 'app/officer/president/sidebar.html',
        'co_adviser': 'app/co_adviser/sidebar.html',
        'adviser': 'app/adviser/sidebar.html',
        "secretary":    "app/officer/secretary/sidebar.html",
    }

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        org = self.get_organization()
        user = self.request.user
        context["base_template"] = self.role_templates.get(user.role, 'app/base.html')
        context["students"] = CustomUser.objects.filter(student__organization=org)
        return context

class BulkImportStudentsView(LoginRequiredMixin, TemplateView):

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get("excel_file")

        if not excel_file:
            messages.error(request, "No file uploaded.")
            return redirect(request.META.get("HTTP_REFERER", "/"))

        if not excel_file.name.endswith(".xlsx"):
            messages.error(request, "Only .xlsx files are supported.")
            return redirect(request.META.get("HTTP_REFERER", "/"))

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
        except Exception:
            messages.error(request, "Invalid Excel file.")
            return redirect(request.META.get("HTTP_REFERER", "/"))

        preview_data = []
        error_rows   = []
        seen_emails      = set()
        seen_student_ids = set()

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        for idx, row in enumerate(rows, start=2):
            if not any(row):
                continue

            try:
                first_name     = str(row[0]).strip() if row[0] is not None else ""
                last_name      = str(row[1]).strip() if row[1] is not None else ""
                email          = str(row[2]).strip() if row[2] is not None else ""
                student_id_raw = row[3]
                program        = str(row[4]).strip() if row[4] is not None else ""
                year_raw       = row[5]
                section        = str(row[6]).strip() if row[6] is not None else ""
                status         = str(row[7]).strip() if row[7] is not None else ""
                
            except IndexError:
                error_rows.append(f"Row {idx}: Not enough columns.")
                continue

            missing = []
            if not first_name:     missing.append("first_name")
            if not last_name:      missing.append("last_name")
            if not email:          missing.append("email")
            if not student_id_raw: missing.append("student_id")
            if not program:        missing.append("program")
            if not year_raw:       missing.append("year")
            if not section:        missing.append("section")
            if not status:         missing.append("status")

            if missing:
                error_rows.append(f"Row {idx}: Missing {', '.join(missing)}")
                continue


            try:
                student_id = int(float(student_id_raw))
            except (ValueError, TypeError):
                error_rows.append(f"Row {idx}: Invalid student_id '{student_id_raw}'")
                continue


            try:
                year = int(float(year_raw))
                if year not in (1, 2, 3, 4):
                    raise ValueError
            except (ValueError, TypeError):
                error_rows.append(f"Row {idx}: Year must be 1-4")
                continue

            valid_programs = [p[0] for p in Student.PROGRAM_CHOICE]
            if program not in valid_programs:
                error_rows.append(f"Row {idx}: Invalid program '{program}'")
                continue

            if CustomUser.objects.filter(email=email).exists():
                error_rows.append(f"Row {idx}: Email already exists '{email}'")
                continue


            if email in seen_emails:
                error_rows.append(f"Row {idx}: Duplicate email in file '{email}'")
                continue

            if Student.objects.filter(student_id=student_id).exists():
                error_rows.append(f"Row {idx}: Student ID already exists '{student_id}'")
                continue

            if student_id in seen_student_ids:
                error_rows.append(f"Row {idx}: Duplicate student_id in file '{student_id}'")
                continue

            seen_emails.add(email)
            seen_student_ids.add(student_id)

            preview_data.append({
                "first_name": first_name,
                "last_name":  last_name,
                "email":      email,
                "student_id": student_id,
                "program":    program,
                "year":       year,
                "section":    section,
                "status":     status,
                
            })

        if not preview_data:
            messages.error(request, "No valid rows found.")
            return redirect(request.META.get("HTTP_REFERER", "/"))

        request.session["import_preview"] = preview_data
        request.session["import_errors"]  = error_rows

        return redirect(request.META.get("HTTP_REFERER", "/"))


class ConfirmImportStudentsView(LoginRequiredMixin, View):

    def post(self, request, *args, **kwargs):
        data = request.session.get("import_preview", [])

        if not data:
            messages.error(request, "No data to import.")
            return redirect("members")
        organization = None
        try:
            organization = request.user.officer.organization
        except AttributeError:
            pass

        created_credentials = []
        created_count       = 0

        try:
            with transaction.atomic():
                for row in data:
                    temp_password = "TransparenSee"

                    user = CustomUser.objects.create_user(
                        username=row["email"],
                        email=row["email"],
                        first_name=row["first_name"],
                        last_name=row["last_name"],
                        password=temp_password,
                        role="student",
                    )

                    Student.objects.create(
                        user=user,
                        student_id=row["student_id"],
                        program=row["program"],
                        year=row["year"],
                        section=row["section"],
                        status=row["status"],
                        organization=organization, 
                    )

                    created_credentials.append({
                        "name":     user.get_full_name(),
                        "email":    row["email"],
                        "password": temp_password,
                    })

                    created_count += 1

        except Exception as exc:
            messages.error(request, f"Import failed: {exc}")
            return redirect("members")

        request.session["import_credentials"] = created_credentials
        request.session.pop("import_preview", None)
        request.session.pop("import_errors",  None)

        messages.success(request, f"{created_count} student(s) imported successfully.")
        return redirect("members")


class DownloadStudentTemplateView(LoginRequiredMixin, TemplateView):
    """Returns a pre-formatted Excel template the president can fill in."""

    def get(self, request, *args, **kwargs):
        import io
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"

        headers = [
            "first_name",
            "last_name",
            "email",
            "student_id",
            "program",
            "year (1-4)",
            "section",
            "status",
        ]

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", start_color="1D4ED8")
        center_align = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, start=1):
            cell           = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align

        sample = ["Juan", "Dela Cruz", "juan.delacruz@school.edu", 20240001, "BSIT", 2, "A", "active"]
        for col_idx, val in enumerate(sample, start=1):
            ws.cell(row=2, column=col_idx, value=val)

        col_widths = [14, 14, 30, 14, 10, 12, 10, 12]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        ws_note = wb.create_sheet("Program Codes")
        ws_note["A1"] = "Valid Program Codes"
        ws_note["A1"].font = Font(bold=True)
        programs = [
            ("BSIT",     "Bachelor of Science in Information Technology"),
            ("BSCS",     "Bachelor of Science in Computer Science"),
            ("BSP",      "Bachelor of Science in Psychology"),
            ("BSED-MTH", "Bachelor of Secondary Education - Mathematics"),
            ("BSED-ENG", "Bachelor of Secondary Education - English"),
            ("BSHM",     "Bachelor of Science in Hospitality Management"),
            ("BSC",      "Bachelor of Science in Criminology"),
            ("BSBA-MM",  "Bachelor of Science in Business Administration - Marketing Management"),
            ("BSBA-HR",  "Bachelor of Science in Business Administration - Human Resource Management"),
        ]
        for row_idx, (code, label) in enumerate(programs, start=2):
            ws_note.cell(row=row_idx, column=1, value=code)
            ws_note.cell(row=row_idx, column=2, value=label)
        ws_note.column_dimensions["A"].width = 12
        ws_note.column_dimensions["B"].width = 65

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="student_import_template.xlsx"'
        return response


class ClearImportCredentialsView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        request.session.pop("import_credentials", None)
        return JsonResponse({"ok": True})


class ClearImportPreviewView(LoginRequiredMixin, View):
    def post(self, request):
        request.session.pop("import_preview", None)
        request.session.pop("import_errors", None)
        return JsonResponse({"ok": True})
    
class LogsView(RoleRequireMixin, ListView):
    role_required = ['treasurer', 'auditor', 'president', 'vice_president', 'co_adviser', 'adviser', 'head', 'campus_admin', 'admin', 'secretary']
    template_name = 'app/logs.html'
    paginate_by = 10
    model = ReportApprovalLog
    context_object_name = 'logs'

    def get_organization(self):
        user = self.request.user
        if hasattr(user, 'officer'):
            return user.officer.organization
        elif hasattr(user, 'adviser'):
            return user.adviser.organization
        elif hasattr(user, 'co_adviser'):
            return user.adviser.organization
        return None
    
    role_templates = {
        'treasurer': 'app/officer/treasurer/sidebar.html',
        'auditor': 'app/officer/auditor/sidebar.html',
        'president': 'app/officer/president/sidebar.html',
        'vice_president': 'app/officer/president/sidebar.html',
        'head': 'app/heads/sidebar.html',
        'campus_admin': 'app/campus_admin/sidebar.html',
        'admin': 'app/superadmin/sidebar.html',
        "secretary":    "app/officer/secretary/sidebar.html",

    }
    
    def get_queryset(self):
        user = self.request.user
        org = self.get_organization()
        if hasattr(user, 'officer'):
            return ReportApprovalLog.objects.filter(report__organization=org)
        return ReportApprovalLog.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        context["base_template"] = self.role_templates.get(user.role, 'app/base.html')
        context['logs_count'] = self.get_queryset().count()

        return context
    
